"""
Receipt generation (PDF + printer simulation) and Excel export utilities.
Uses Qt's own QPrinter so no extra PDF dependency is required.
"""
from __future__ import annotations

import os
from typing import Optional

from PySide6.QtCore import QSizeF, QMarginsF, Qt, QRectF
from PySide6.QtGui import QPainter, QFont, QPageSize, QPageLayout
from PySide6.QtPrintSupport import QPrinter, QPrintDialog
from PySide6.QtWidgets import QWidget

from shared import db


def format_currency(amount: float) -> str:
    try:
        amount = float(amount)
    except (TypeError, ValueError):
        amount = 0.0
    negative = amount < 0
    amount = abs(amount)
    whole = int(amount)
    frac = round((amount - whole) * 100)
    s = str(whole)
    if len(s) > 3:
        last3 = s[-3:]
        rest = s[:-3]
        groups = []
        while len(rest) > 2:
            groups.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            groups.insert(0, rest)
        s = ",".join(groups) + "," + last3
    return f"{'-' if negative else ''}₹{s}.{frac:02d}"


def _build_receipt_page(printer: QPrinter, atm_name: str, atm_code: str, lines: list[tuple[str, str]],
                         title: str = "TRANSACTION RECEIPT"):
    painter = QPainter(printer)
    width = printer.width()
    y = 40
    left = 20

    def draw_center(text, font_size=13, bold=True, dy=26):
        nonlocal y
        f = QFont("Consolas", font_size)
        f.setBold(bold)
        painter.setFont(f)
        painter.drawText(QRectF(0, y, width, dy + 10), Qt.AlignmentFlag.AlignHCenter, text)
        y += dy

    def draw_line(dy=18):
        nonlocal y
        painter.drawLine(left, y, width - left, y)
        y += dy

    def draw_row(label, value, dy=24):
        nonlocal y
        f = QFont("Consolas", 10)
        painter.setFont(f)
        painter.drawText(QRectF(left, y, width * 0.55, dy), Qt.AlignmentFlag.AlignLeft, label)
        painter.drawText(QRectF(width * 0.5, y, width * 0.5 - left, dy),
                          Qt.AlignmentFlag.AlignRight, value)
        y += dy

    draw_center("SimBank ATM Network", 15, True, 30)
    draw_center("*** SIMULATION / TEST ONLY ***", 9, True, 20)
    draw_center(title, 12, True, 26)
    draw_center(f"{atm_name} ({atm_code})", 9, False, 20)
    draw_line(16)
    for label, value in lines:
        draw_row(label, value)
    draw_line(16)
    draw_center("No real funds were moved.", 8, False, 16)
    draw_center("Thank you for banking with SimBank", 9, True, 20)
    painter.end()


def generate_receipt_pdf(out_path: str, atm_name: str, atm_code: str,
                          lines: list[tuple[str, str]], title: str = "TRANSACTION RECEIPT") -> str:
    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    printer.setOutputFileName(out_path)
    page_size = QPageSize(QSizeF(80, 200), QPageSize.Unit.Millimeter, "Receipt")
    printer.setPageSize(page_size)
    printer.setPageMargins(QMarginsF(4, 4, 4, 4), QPageLayout.Unit.Millimeter)
    _build_receipt_page(printer, atm_name, atm_code, lines, title)
    return out_path


def print_receipt_dialog(parent: Optional[QWidget], atm_name: str, atm_code: str,
                          lines: list[tuple[str, str]], title: str = "TRANSACTION RECEIPT") -> bool:
    """Opens the OS print dialog and prints (or prints to a virtual/PDF printer)."""
    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    page_size = QPageSize(QSizeF(80, 200), QPageSize.Unit.Millimeter, "Receipt")
    printer.setPageSize(page_size)
    dialog = QPrintDialog(printer, parent)
    dialog.setWindowTitle("Print Simulated ATM Receipt")
    if dialog.exec() == QPrintDialog.DialogCode.Accepted:
        _build_receipt_page(printer, atm_name, atm_code, lines, title)
        return True
    return False


def txn_receipt_lines(txn_row, extra: Optional[dict] = None) -> list[tuple[str, str]]:
    lines = [
        ("Txn ID", txn_row["txn_id"]),
        ("Date/Time", txn_row["timestamp"]),
        ("Type", txn_row["txn_type"]),
        ("Status", txn_row["status"]),
        ("Amount", format_currency(txn_row["amount"])),
    ]
    if txn_row["balance_after"] is not None:
        lines.append(("Balance", format_currency(txn_row["balance_after"])))
    if txn_row["related_account"]:
        lines.append(("To Account", txn_row["related_account"]))
    if txn_row["remarks"]:
        lines.append(("Remarks", txn_row["remarks"]))
    if extra:
        for k, v in extra.items():
            lines.append((k, str(v)))
    return lines


def export_transactions_excel(out_path: str, rows) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    headers = ["Txn ID", "Date/Time", "Type", "Amount", "Balance After",
               "Related Account", "ATM", "Status", "Remarks"]
    ws.append(headers)
    header_fill = PatternFill(start_color="121C2E", end_color="121C2E", fill_type="solid")
    header_font = Font(color="00C896", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            r["txn_id"], r["timestamp"], r["txn_type"], r["amount"], r["balance_after"],
            r["related_account"], r["atm_code"], r["status"], r["remarks"],
        ])

    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(header) + 4)

    ws.freeze_panes = "A2"
    wb.save(out_path)
    return out_path
